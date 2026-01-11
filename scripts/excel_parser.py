"""
Excel Parser for Rooster Schedule

Parses the Excel schedule file and extracts employee schedules dynamically.
Handles new employees, removed employees, and various edge cases.
"""

import os
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class ShiftEntry:
    """Represents a single shift entry for an employee."""
    date: datetime
    code: str


@dataclass
class Employee:
    """Represents an employee with their schedule."""
    name: str
    group: str
    shifts: List[ShiftEntry]

    @property
    def filename(self) -> str:
        """Generate a safe filename for this employee."""
        # Normalize name: lowercase, replace spaces with underscores
        safe_name = self.name.lower().strip()
        safe_name = re.sub(r'\s+', '_', safe_name)
        # Remove special characters except underscores
        safe_name = re.sub(r'[^a-z0-9_]', '', safe_name)
        return f"{safe_name}.ics"


class ExcelParser:
    """Parses the Rooster Excel file to extract employee schedules."""

    def __init__(self, config: dict):
        """
        Initialize the parser with configuration.

        Args:
            config: Configuration dictionary with Excel structure settings
        """
        self.config = config
        self.excel_config = config.get('excel', {})

        # Excel structure settings
        self.date_row = self.excel_config.get('date_row', 5)
        self.first_employee_row = self.excel_config.get('first_employee_row', 7)
        self.name_column = self.excel_config.get('name_column', 3)
        self.group_column = self.excel_config.get('group_column', 2)
        self.first_schedule_column = self.excel_config.get('first_schedule_column', 4)

        # Valid shift codes from config
        self.timed_shifts = set(config.get('timed_shifts', {}).keys())
        self.allday_events = set(config.get('allday_events', {}).keys())
        self.ignore_codes = set(config.get('ignore_codes', ['', ' ']))
        self.valid_codes = self.timed_shifts | self.allday_events

    def parse(self, excel_path: str) -> Dict[str, Employee]:
        """
        Parse the Excel file and return employee schedules.

        Args:
            excel_path: Path to the Excel file

        Returns:
            Dictionary mapping employee filename to Employee object
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Load workbook (data_only=True to get values, not formulas)
        wb = load_workbook(excel_path, data_only=True)

        # Use the first sheet (main schedule)
        ws = wb.active

        # Parse date headers
        dates = self._parse_dates(ws)

        if not dates:
            raise ValueError("No valid dates found in the Excel file")

        # Parse employees and their schedules
        employees = self._parse_employees(ws, dates)

        wb.close()

        return employees

    def _parse_dates(self, ws) -> Dict[int, datetime]:
        """
        Parse date headers from the worksheet.

        Args:
            ws: Worksheet object

        Returns:
            Dictionary mapping column index to date
        """
        dates = {}

        # Find the maximum column with data
        max_col = ws.max_column

        for col in range(self.first_schedule_column, max_col + 1):
            cell_value = ws.cell(row=self.date_row, column=col).value

            if cell_value is None:
                continue

            # Handle different date formats
            if isinstance(cell_value, datetime):
                dates[col] = cell_value
            elif isinstance(cell_value, (int, float)):
                # Excel serial date number
                try:
                    # Excel epoch is December 30, 1899
                    from datetime import timedelta
                    excel_epoch = datetime(1899, 12, 30)
                    dates[col] = excel_epoch + timedelta(days=int(cell_value))
                except (ValueError, OverflowError):
                    continue
            elif isinstance(cell_value, str):
                # Try to parse string date
                for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        dates[col] = datetime.strptime(cell_value, fmt)
                        break
                    except ValueError:
                        continue

        return dates

    def _parse_employees(self, ws, dates: Dict[int, datetime]) -> Dict[str, Employee]:
        """
        Parse employee data from the worksheet.

        Args:
            ws: Worksheet object
            dates: Dictionary mapping column index to date

        Returns:
            Dictionary mapping employee filename to Employee object
        """
        employees = {}
        current_group = ""

        max_row = ws.max_row

        for row in range(self.first_employee_row, max_row + 1):
            # Get group/role from column B
            group_cell = ws.cell(row=row, column=self.group_column).value
            if group_cell:
                group_str = str(group_cell).strip()
                # Check if this is a group header (like "PLOEG A")
                if group_str.upper().startswith('PLOEG') or group_str.upper() in [
                    'SAMPLE RETAIN ASSISTANTS', 'DAY TIME SUPPORT',
                    'INTERNS', '0H CONTRACTS'
                ]:
                    current_group = group_str
                    continue

            # Get employee name from column C
            name_cell = ws.cell(row=row, column=self.name_column).value

            if not name_cell:
                continue

            name = str(name_cell).strip()

            # Skip header rows or empty names
            if not name or name.upper() in ['NAME', 'NAAM', 'EMPLOYEE']:
                continue

            # Skip if name looks like a group header
            if name.upper().startswith('PLOEG'):
                current_group = name
                continue

            # Parse shifts for this employee
            shifts = []
            for col, date in dates.items():
                cell_value = ws.cell(row=row, column=col).value

                if cell_value is None:
                    continue

                code = str(cell_value).strip().upper()

                # Skip ignored codes
                if code in self.ignore_codes or not code:
                    continue

                # Only include known shift codes
                if code in self.valid_codes:
                    shifts.append(ShiftEntry(date=date, code=code))

            # Only add employees who have at least one shift
            if shifts:
                employee = Employee(
                    name=name,
                    group=current_group,
                    shifts=shifts
                )

                # Handle duplicate names by appending group
                filename = employee.filename
                if filename in employees:
                    # Add group to make unique
                    safe_group = re.sub(r'[^a-z0-9]', '_', current_group.lower())
                    filename = f"{filename[:-4]}_{safe_group}.ics"
                    employee = Employee(
                        name=f"{name} ({current_group})",
                        group=current_group,
                        shifts=shifts
                    )

                employees[filename] = employee

        return employees


def parse_excel(excel_path: str, config: dict) -> Dict[str, Employee]:
    """
    Convenience function to parse an Excel file.

    Args:
        excel_path: Path to the Excel file
        config: Configuration dictionary

    Returns:
        Dictionary mapping employee filename to Employee object
    """
    parser = ExcelParser(config)
    return parser.parse(excel_path)
